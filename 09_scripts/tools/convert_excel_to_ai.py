"""
엑셀(.xlsx)을 AI가 읽기 쉬운 형태(JSON / JSONL / Markdown)로 변환합니다.

- 각 시트의 첫 행(header)을 속성명으로 사용합니다.
- 셀 값의 개행/공백을 정규화하고, 완전히 빈 행은 제외합니다.
- 출력은 JSONL, JSON, (선택) Markdown 세 형태로 생성합니다.
- 시트가 여러 개일 때만 결과 파일명에 '_시트명' 접미사를 붙입니다(시트 1개면 접미사 없음).

이 스크립트는 09_scripts/tools/ 에 있으며, 보통 런처(09_scripts/run.ps1)로 실행합니다.

권장 사용 흐름:
  입력 엑셀(01_inputs/)을 AI 친화적 데이터(03_working/data/)로 변환합니다.
    python 09_scripts/tools/convert_excel_to_ai.py 01_inputs/samples/data.xlsx -o 03_working/data

사용 예:
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx -o 03_working/data
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx --sheet Sheet1
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx --all-sheets
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx --no-md
    python 09_scripts/tools/convert_excel_to_ai.py input.xlsx --list-sheets

주의:
- 실데이터/개인정보/접속정보가 담긴 엑셀은 SECURITY_RULES.md를 따릅니다.
  민감 자료의 변환 결과는 Git에 커밋하지 않습니다.
"""

from pathlib import Path
import json
import sys
import argparse

# Windows 콘솔 등에서 한글 진행 메시지가 깨지지 않도록 stdout을 UTF-8로 설정
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

try:
    import openpyxl
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다.")
    print("설치 명령: pip install openpyxl")
    sys.exit(1)


def normalize_value(value):
    if value is None:
        return ""

    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    lines = [line.rstrip() for line in text.split("\n")]
    return "\n".join(lines).strip()


def normalize_header(value):
    header = normalize_value(value)
    return header.strip()


def sheet_to_rows(ws):
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [normalize_header(v) for v in rows[0]]

    if not any(headers):
        return []

    result = []

    for row in rows[1:]:
        item = {}

        for header, value in zip(headers, row):
            if not header:
                continue
            item[header] = normalize_value(value)

        # 완전히 빈 행 제외
        if not any(v for v in item.values()):
            continue

        result.append(item)

    return result


def write_jsonl(rows, output_path):
    with output_path.open("w", encoding="utf-8") as f:
        for item in rows:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def write_json(rows, output_path):
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def write_md(rows, output_path, title):
    lines = [f"# {title}", ""]

    for idx, item in enumerate(rows, start=1):
        first_value = next((v for v in item.values() if v), f"ROW-{idx}")

        lines.append(f"## {idx}. {first_value}")
        lines.append("")

        for key, value in item.items():
            lines.append(f"### {key}")
            lines.append("")
            lines.append(value if value else "")
            lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(
        description="엑셀 첫 행을 속성명으로 사용하여 JSON/JSONL/Markdown으로 변환합니다."
    )

    parser.add_argument(
        "input",
        help="변환할 xlsx 파일 경로"
    )

    parser.add_argument(
        "-o",
        "--output-dir",
        default=None,
        help="출력 폴더. 생략하면 입력 파일과 같은 폴더에 생성"
    )

    parser.add_argument(
        "--sheet",
        default=None,
        help="특정 시트명만 변환. 생략하면 첫 번째 시트만 변환"
    )

    parser.add_argument(
        "--all-sheets",
        action="store_true",
        help="모든 시트를 각각 변환"
    )

    parser.add_argument(
        "--no-md",
        action="store_true",
        help="Markdown 파일은 생성하지 않음"
    )

    parser.add_argument(
        "--list-sheets",
        action="store_true",
        help="변환하지 않고 시트 이름 목록만 출력(한 줄에 하나)하고 종료"
    )

    args = parser.parse_args()

    input_path = Path(args.input).resolve()

    if not input_path.exists():
        print(f"파일을 찾을 수 없습니다: {input_path}")
        sys.exit(1)

    if input_path.suffix.lower() != ".xlsx":
        print("현재는 .xlsx 파일만 지원합니다.")
        sys.exit(1)

    output_dir = Path(args.output_dir).resolve() if args.output_dir else input_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(input_path, data_only=True)

    # 시트 이름만 출력하고 종료 (런처 등에서 시트 개수 판단용)
    if args.list_sheets:
        for name in wb.sheetnames:
            print(name)
        return

    if args.all_sheets:
        target_sheets = wb.worksheets
    elif args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"시트를 찾을 수 없습니다: {args.sheet}")
            print(f"사용 가능한 시트: {wb.sheetnames}")
            sys.exit(1)
        target_sheets = [wb[args.sheet]]
    else:
        target_sheets = [wb.active]

    total_count = 0

    for ws in target_sheets:
        rows = sheet_to_rows(ws)
        total_count += len(rows)

        safe_sheet_name = "".join(
            c if c not in r'\/:*?"<>|' else "_"
            for c in ws.title
        )

        # 시트명 접미사는 '실제로 여러 시트를 변환할 때만' 붙인다.
        # (--all-sheets 여도 시트가 1개뿐이면 접미사 없음)
        if len(target_sheets) > 1:
            base_name = f"{input_path.stem}_{safe_sheet_name}"
        else:
            base_name = input_path.stem

        jsonl_path = output_dir / f"{base_name}.jsonl"
        json_path = output_dir / f"{base_name}.json"
        md_path = output_dir / f"{base_name}.md"

        write_jsonl(rows, jsonl_path)
        write_json(rows, json_path)

        if not args.no_md:
            write_md(rows, md_path, title=base_name)

        print(f"[{ws.title}] 변환 완료")
        print(f"- 행 수: {len(rows)}")
        print(f"- JSONL: {jsonl_path}")
        print(f"- JSON : {json_path}")

        if not args.no_md:
            print(f"- MD   : {md_path}")

    print("")
    print(f"전체 완료. 총 변환 행 수: {total_count}")


if __name__ == "__main__":
    main()
